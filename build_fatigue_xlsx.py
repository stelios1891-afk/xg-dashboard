import json, pandas as pd, numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def isodate(s):
    try: return datetime.strptime(s.replace(' UTC',''),'%a, %b %d, %Y, %H:%M').strftime('%Y-%m-%d')
    except: return s[:10]
TOP5=['EPL','LaLiga','SerieA','Bundesliga','Ligue1']
ATT,DEF=0.89,1.18
events=[]
for lg in TOP5:
    for sea in ['2425','2526']:
        d=json.load(open(f'data_{lg}_{sea}.json'))
        for m in d.values():
            if m['hs'] is None: continue
            dt=isodate(m['date']); events.append((dt,int(m['home']['id']),'H','LEAGUE')); events.append((dt,int(m['away']['id']),'A','LEAGUE'))
for e in json.load(open('euro_fixtures.json')):
    events.append((e['date'],e['home'],'H',e['comp'])); events.append((e['date'],e['away'],'A',e['comp']))
tl={}
for dt,t,v,c in events: tl.setdefault(t,[]).append((dt,v,c))
for t in tl: tl[t]=sorted(set(tl[t]),key=lambda x:x[0])
def tired(team,date,ven):
    pm=[x for x in tl.get(team,[]) if x[0]<date][-3:]
    venues=[v for _,v,_ in pm[-2:]]+[ven]
    ok=len(venues)==3 and all(v=='A' for v in venues) and len(pm)>=1 and pm[-1][2] in ('CL','EL','ECL') and pm[-1][1]=='A'
    return ok,(pm[-1] if pm else None)
import unicodedata,re
def norm(s):
    s=unicodedata.normalize('NFKD',str(s)).encode('ascii','ignore').decode().lower()
    return frozenset(set(re.sub(r'[^a-z ]',' ',s).split())-{'fc','cf','ac','calcio','club','de','afc','ss','us','as','rc','ud','sd','sc','rcd','real','1','vfl','vfb','tsg','sv','bsc','og'})
ALIAS={'Athletic Club':'Ath Bilbao','Borussia Mönchengladbach':"M'gladbach",'Espanyol':'Espanol','Hamburger SV':'Hamburg','Wolverhampton Wanderers':'Wolves'}
O=[]
for lg in TOP5:
    for sea in ['2425','2526']:
        o=pd.read_csv(f'odds/{lg}_{sea}.csv',encoding='latin-1'); o['season']=sea; O.append(o)
O=pd.concat(O,ignore_index=True)
fdn={}
for _,r in O.iterrows(): fdn[r['HomeTeam']]=norm(r['HomeTeam']); fdn[r['AwayTeam']]=norm(r['AwayTeam'])
def resolve(n):
    tn=norm(ALIAS.get(n,n)); best=None;bs=0;bj=0
    for raw,tg in fdn.items():
        ov=len(tn&tg)
        if ov>bs or (ov==bs and ov/max(len(tn|tg),1)>bj): bs=ov;bj=ov/max(len(tn|tg),1);best=raw
    return best if bs>0 else None
Om={}
for _,r in O.iterrows(): Om[(str(r['season']),r['HomeTeam'],r['AwayTeam'])]=r
def settle(g,side,line,odds):
    parts=[line] if (line*4)%2==0 else [line-0.25,line+0.25]; pnl=0.
    for L in parts:
        s=1/len(parts); m=(g if side==1 else -g)+L
        if m>0.01: pnl+=s*(odds-1)
        elif abs(m)<0.01: pnl+=0.
        else: pnl-=s
    return pnl

P=pd.read_csv('model_predictions.csv'); P=P[P.league.isin(TOP5)].copy()
P['season']=P['season'].astype(str)
P=P.sort_values(['season','date'])
rows=[]
for _,r in P.iterrows():
    for tn_,on_,tid,ven in [(r['home_name'],r['away_name'],int(r['home']),'H'),(r['away_name'],r['home_name'],int(r['away']),'A')]:
        ok,prev=tired(tid,r['date'],ven)
        if not ok: continue
        xh,xa=r['xg_h'],r['xg_a']; th=(ven=='H')
        axh,axa=(xh*ATT,xa*DEF) if th else (xh*DEF,xa*ATT)
        o=Om.get((r['season'],resolve(r['home_name']),resolve(r['away_name'])))
        line=oh=oa=None
        if o is not None:
            line=o.get('AHCh')
            for hh,ha in [('BFECAHH','BFECAHA'),('PCAHH','PCAHA'),('AvgCAHH','AvgCAHA')]:
                if pd.notna(o.get(hh)) and pd.notna(o.get(ha)): oh,oa=float(o[hh]),float(o[ha]); break
        pnl=fodds=fpick=None
        if line is not None and oh is not None and pd.notna(line):
            line=float(line)
            if th: fpick,side,L,fodds=r['away_name'],-1,-line,oa
            else:  fpick,side,L,fodds=r['home_name'],1,line,oh
            pnl=settle(r['gd'],side,L,fodds)
        rows.append([f"20{r['season'][:2]}/{r['season'][2:]}",r['date'],r['league'],
            r['home_name'],r['away_name'],tn_,(f"{prev[2]} {prev[0]}" if prev else ''),
            round(xh,2),round(xa,2),round(xh-xa,2),round(axh,2),round(axa,2),round(axh-axa,2),
            (line if line is not None else ''),(oh if oh else ''),(oa if oa else ''),
            f"{r['hg']}-{r['ag']}",int(r['gd']),(fpick or ''),(fodds or ''),
            (('WIN' if pnl>0 else ('PUSH' if abs(pnl)<0.01 else 'LOSS')) if pnl is not None else ''),
            (round(pnl,3) if pnl is not None else '')])

cols=['Σεζόν','Ημ/νία','Λίγκα','Γηπεδούχος','Φιλοξ.','Κουρασμένη','Προηγ. Ευρωπαϊκό',
      'xG γηπ (πριν)','xG φιλ (πριν)','Supr. πριν','xG γηπ (μετά)','xG φιλ (μετά)','Supr. μετά',
      'AH γραμμή','Odds γηπ','Odds φιλ','Σκορ','GD','Ποντάρουμε','Fade odds','Fade','Fade P&L (u)']
wb=Workbook(); ws=wb.active; ws.title='Κουρασμένα ματς'
hdr_fill=PatternFill('solid',fgColor='1F3864'); hdr_font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
grp_fill=PatternFill('solid',fgColor='2E5496')  # projections μετα
thin=Side(style='thin',color='D9D9D9'); border=Border(thin,thin,thin,thin)
for j,c in enumerate(cols,1):
    cell=ws.cell(1,j,c); cell.fill=hdr_fill; cell.font=hdr_font; cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True); cell.border=border
win_f=PatternFill('solid',fgColor='C6EFCE'); loss_f=PatternFill('solid',fgColor='FFC7CE'); push_f=PatternFill('solid',fgColor='FFEB9C')
adj_f=PatternFill('solid',fgColor='EAF0FA')
for i,row in enumerate(rows,2):
    for j,val in enumerate(row,1):
        cell=ws.cell(i,j,val); cell.font=Font(name='Arial',size=10); cell.border=border
        cell.alignment=Alignment(horizontal='center' if j not in (4,5,6,7,19) else 'left')
        if j in (11,12,13): cell.fill=adj_f
        if j==21:  # Fade result color
            if val=='WIN': cell.fill=win_f
            elif val=='LOSS': cell.fill=loss_f
            elif val=='PUSH': cell.fill=push_f
# summary row
last=len(rows)+2
ws.cell(last+1,1,'ΣΥΝΟΛΟ FADE:').font=Font(name='Arial',bold=True,size=10)
ws.cell(last+1,20).value=f'=COUNTIF(U2:U{last-1},"WIN")&"W / "&COUNTIF(U2:U{last-1},"LOSS")&"L"'
ws.cell(last+1,20).font=Font(name='Arial',bold=True,size=10)
ws.cell(last+1,22).value=f'=SUM(V2:V{last-1})'
ws.cell(last+1,22).font=Font(name='Arial',bold=True,size=10); ws.cell(last+1,22).number_format='0.00'
ws.cell(last+2,1,'ROI Fade:').font=Font(name='Arial',bold=True,size=10)
ws.cell(last+2,22).value=f'=V{last+1}/COUNTA(V2:V{last-1})'
ws.cell(last+2,22).number_format='0.0%'; ws.cell(last+2,22).font=Font(name='Arial',bold=True,size=10)
# widths
widths=[9,11,10,18,18,16,16,11,11,9,11,11,9,8,7,7,7,5,18,8,7,10]
for j,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(j)].width=w
ws.freeze_panes='D2'
ws.auto_filter.ref=f"A1:V{last-1}"
# legend sheet
ws2=wb.create_sheet('Επεξήγηση')
notes=[['ΣΤΗΛΕΣ',''],
 ['Κουρασμένη','Ομάδα στο 3ο διαδοχικό εκτός με ευρωπαϊκό ταξίδι στη μέση'],
 ['xG (πριν)','Τα projections του μοντέλου μας ΧΩΡΙΣ διόρθωση κόπωσης'],
 ['xG (μετά)','ΜΕ adjustment: κουρασμένη επίθεση ×0.89, κουρασμένη άμυνα ×1.18 (αντίπαλος επίθεση)'],
 ['AH γραμμή / Odds','Betfair Exchange closing Asian Handicap (fallback Pinnacle/Avg)'],
 ['Ποντάρουμε','Ο αντίπαλος της κουρασμένης (fade) — στη γραμμή/odds της αγοράς'],
 ['Fade P&L (u)','Κέρδος/ζημιά σε units του 1 (flat stake). +0.9 = κερδίσαμε 0.9× τη μίζα'],
 ['',''],
 ['ΠΡΟΣΟΧΗ',''],
 ['Δείγμα','36 ματς σε 2 σεζόν — στατιστικά ασθενές (~2.2 SE). Ενδεικτικό, όχι αποδεδειγμένο.'],
 ['Έδρα','Και τα 36 η κουρασμένη παίζει εκτός· μέρος του edge μπορεί να ειναι απλως πλεονεκτημα εδρας (ανελεγκτο)'],
 ['Adjustment','Τα ×0.89/×1.18 προερχονται απο τα ιδια 36 ματς (in-sample) — μην τα κλειδωσεις ως ακριβεις τιμες']]
for i,(a,b) in enumerate(notes,1):
    ws2.cell(i,1,a).font=Font(name='Arial',bold=(b=='' or a in('ΣΤΗΛΕΣ','ΠΡΟΣΟΧΗ')),size=10)
    ws2.cell(i,2,b).font=Font(name='Arial',size=10); ws2.cell(i,2).alignment=Alignment(wrap_text=True)
ws2.column_dimensions['A'].width=16; ws2.column_dimensions['B'].width=80
wb.save('/mnt/user-data/outputs/fatigue_matches.xlsx')
print(f"Εγινε: {len(rows)} ματς")
